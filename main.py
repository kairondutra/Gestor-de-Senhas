import os
import openpyxl
from tkinter import messagebox
import customtkinter as ctk
import json
import pyautogui
import time

# Configuração e banco de dados
CONFIG_PATH = os.path.join("src", "config.json")
DB_PATH = os.path.join("src", "db.xlsx")

# Carregar ou salvar configurações
def load_config():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r') as file:
                return json.load(file)
        except (IOError, json.JSONDecodeError):
            pass
    return {"account_coords": None, "password_coords": None, "theme": "Dark"}

def save_config(config):
    os.makedirs(os.path.dirname(CONFIG_PATH), exist_ok=True)
    with open(CONFIG_PATH, 'w') as file:
        json.dump(config, file)

# Verifica ou cria o banco de dados
def initialize_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    if not os.path.exists(DB_PATH):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Accounts"
        sheet.append(["Account Name", "Username", "Password"])
        workbook.save(DB_PATH)

def read_accounts():
    initialize_db()
    try:
        workbook = openpyxl.load_workbook(DB_PATH)
        sheet = workbook.active
        accounts = [
            {"name": row[0], "username": row[1], "password": row[2]}
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if all(row)
        ]
        return accounts
    except FileNotFoundError:
        return []

def save_account(account_name, username, password):
    if not all([account_name, username, password]):
        raise ValueError("Todos os campos devem ser preenchidos.")
    workbook = openpyxl.load_workbook(DB_PATH)
    sheet = workbook.active
    sheet.append([account_name, username, password])
    workbook.save(DB_PATH)

def update_account(old_name, new_name, new_username, new_password):
    workbook = openpyxl.load_workbook(DB_PATH)
    sheet = workbook.active
    account_updated = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value and row[0].value.strip().lower() == old_name.strip().lower():
            row[0].value = new_name.strip()
            row[1].value = new_username.strip()
            row[2].value = new_password.strip()
            account_updated = True
            break
    if account_updated:
        workbook.save(DB_PATH)
    else:
        raise ValueError("Conta não encontrada.")

def delete_account(account_name):
    workbook = openpyxl.load_workbook(DB_PATH)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == account_name:
            sheet.delete_rows(row[0].row)
            break
    workbook.save(DB_PATH)

def capture_coordinates():
    config = load_config()

    def capture(type_):
        messagebox.showinfo(
            "Captura de Coordenadas",
            f"Posicione o cursor no campo {type_} e pressione Enter."
        )
        x, y = pyautogui.position()
        config[f"{type_}_coords"] = (x, y)
        save_config(config)
        messagebox.showinfo("Sucesso", f"{type_.capitalize()} coordenadas salvas.")

    capture("account")
    capture("password")

# Interface do usuário
class PasswordManagerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Gestor de Login")
        self.geometry("400x280")
        self.resizable(False, False)

        # Carregar configurações e aplicar tema
        self.config = load_config()
        ctk.set_appearance_mode(self.config.get("theme", "Dark"))

        # Configuração das abas
        self.tab_view = ctk.CTkTabview(self)
        self.tab_view.pack(expand=True, fill="both")

        # Adicionar abas
        self.tab_accounts = self.tab_view.add("Contas")
        self.tab_settings = self.tab_view.add("Configurações")	

        # Atributo para rastrear a página ativa
        self.current_page = 0
        self.accounts_per_page = 12

        # Configurar o conteúdo das abas
        self.setup_accounts_tab()
        self.setup_settings_tab()

    def setup_accounts_tab(self):
        for widget in self.tab_accounts.winfo_children():
            widget.destroy()

        accounts = read_accounts()
        total_pages = (len(accounts) + self.accounts_per_page - 1) // self.accounts_per_page
        start_index = self.current_page * self.accounts_per_page
        end_index = start_index + self.accounts_per_page
        accounts_to_display = accounts[start_index:end_index]

        if accounts_to_display:
            # Grade de botões para exibir cada conta
            grid_frame = ctk.CTkFrame(self.tab_accounts)
            grid_frame.pack(expand=True, fill="both", pady=5)

            for i, account in enumerate(accounts_to_display):
                # Nome da conta no botão
                button = ctk.CTkButton(
                    grid_frame,
                    text=account["name"],
                    command=lambda acc=account: self.perform_login(acc),
                    width=120,
                    height=30,
                )
                button.grid(row=i // 3, column=i % 3, padx=5, pady=5)  # 3 botões por linha

            # Navegação de páginas
            nav_frame = ctk.CTkFrame(self.tab_accounts)
            nav_frame.pack()#fill="x")

            if self.current_page > 0:
                prev_button = ctk.CTkButton(
                    nav_frame, text="<", command=self.prev_page
                )
                prev_button.pack(side="left", padx=10)

            if self.current_page < total_pages - 1:
                next_button = ctk.CTkButton(
                    nav_frame, text=">", command=self.next_page
                )
                next_button.pack(side="right", padx=10)
        else:
            ctk.CTkLabel(
                self.tab_accounts, text="Nenhuma conta encontrada."
            ).pack(pady=10)

    def next_page(self):
        self.current_page += 1
        self.setup_accounts_tab()

    def prev_page(self):
        self.current_page -= 1
        self.setup_accounts_tab()

    def add_account(self):
        self.open_account_editor()

    def open_account_editor(self, account=None):
        popup = ctk.CTkToplevel(self)
        popup.title("Adicionar Conta")
        popup.geometry("300x300")
        popup.resizable(False, False)

        # Variáveis para os campos
        account_name_var = ctk.StringVar()
        username_var = ctk.StringVar()
        password_var = ctk.StringVar()

        if account:
            account_name_var.set(account["name"])
            username_var.set(account["username"])
            password_var.set(account["password"])

        # Layout da janela
        ctk.CTkLabel(popup, text="Nome da Conta").pack(pady=5)
        account_name_entry = ctk.CTkEntry(popup, textvariable=account_name_var)
        account_name_entry.pack(pady=5)

        ctk.CTkLabel(popup, text="Usuário").pack(pady=5)
        username_entry = ctk.CTkEntry(popup, textvariable=username_var)
        username_entry.pack(pady=5)

        ctk.CTkLabel(popup, text="Senha").pack(pady=5)
        password_entry = ctk.CTkEntry(popup, textvariable=password_var, show="*")
        password_entry.pack(pady=5)

        def save_account_callback():
            account_name = account_name_var.get()
            username = username_var.get()
            password = password_var.get()

            if not all([account_name, username, password]):
                messagebox.showerror("Erro", "Preencha todos os campos.")
                return

            if account:
                update_account(account["name"], account_name, username, password)
                messagebox.showinfo("Sucesso", "Conta editada com sucesso!")
            else:
                save_account(account_name, username, password)
                messagebox.showinfo("Sucesso", "Conta adicionada com sucesso!")

            popup.destroy()
            self.setup_accounts_tab()

        ctk.CTkButton(popup, text="Salvar", command=save_account_callback).pack(pady=10)
        popup.protocol("WM_DELETE_WINDOW", popup.destroy)

    def edit_accounts_interface(self):
        popup = ctk.CTkToplevel(self)
        popup.title("Editar Contas")
        popup.geometry("300x250")
        popup.resizable(False, False)

        accounts = read_accounts()
        account_names = [account["name"] for account in accounts]

        # Variáveis
        account_name_var = ctk.StringVar()
        username_var = ctk.StringVar()
        password_var = ctk.StringVar()
        selected_account = {"name": None, "username": None, "password": None}

        def on_account_select(event):
            selected_name = dropdown_var.get()
            for account in accounts:
                if account["name"] == selected_name:
                    selected_account.update(account)
                    account_name_var.set(account["name"])
                    username_var.set(account["username"])
                    password_var.set(account["password"])

        # Dropdown para selecionar a conta
        dropdown_var = ctk.StringVar(value="Selecione a conta")
        dropdown = ctk.CTkOptionMenu(
            popup,
            variable=dropdown_var,
            values=account_names,
            command=on_account_select
        )
        dropdown.pack(pady=10)

        # Campos de edição
        edit_frame = ctk.CTkFrame(popup)
        edit_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(edit_frame, text="Nome da Conta").grid(row=0, column=0, padx=5, pady=5)
        account_name_entry = ctk.CTkEntry(edit_frame, textvariable=account_name_var)
        account_name_entry.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(edit_frame, text="Usuário").grid(row=1, column=0, padx=5, pady=5)
        username_entry = ctk.CTkEntry(edit_frame, textvariable=username_var)
        username_entry.grid(row=1, column=1, padx=5, pady=5)

        ctk.CTkLabel(edit_frame, text="Senha").grid(row=2, column=0, padx=5, pady=5)
        password_entry = ctk.CTkEntry(edit_frame, textvariable=password_var, show="*")
        password_entry.grid(row=2, column=1, padx=5, pady=5)

        def save_changes():
            old_name = selected_account["name"]
            new_name = account_name_var.get().strip()
            new_username = username_var.get().strip()
            new_password = password_var.get().strip()

            if not all([new_name, new_username, new_password]):
                messagebox.showerror("Erro", "Preencha todos os campos.")
                return

            if not old_name:
                messagebox.showerror("Erro", "Selecione uma conta.")
                return

            try:
                update_account(old_name, new_name, new_username, new_password)
                messagebox.showinfo("Sucesso", "Conta editada com sucesso!")
                self.setup_accounts_tab()
                popup.destroy()
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível salvar as alterações: {e}")

        save_button = ctk.CTkButton(edit_frame, text="Salvar", command=save_changes)
        save_button.grid(row=3, column=0, columnspan=2, pady=10)

        popup.protocol("WM_DELETE_WINDOW", popup.destroy)

    def delete_account_interface(self):
        popup = ctk.CTkToplevel(self)
        popup.title("Excluir Conta")
        popup.geometry("300x100")
        popup.resizable(False, False)

        accounts = read_accounts()
        account_names = [account["name"] for account in accounts]

        dropdown_var = ctk.StringVar(value="Selecione uma conta")

        if not accounts:
            messagebox.showerror("Erro", "Nenhuma conta disponível para excluir.")
            popup.destroy()
            return

        dropdown = ctk.CTkOptionMenu(
            popup,
            variable=dropdown_var,
            values=account_names
        )
        dropdown.pack(pady=10)

        def delete_selected():
            selected_name = dropdown_var.get()

            if not selected_name or selected_name == "Selecione uma conta":
                messagebox.showerror("Erro", "Selecione uma conta para excluir.")
                return

            confirm = messagebox.askyesno(
                "Confirmação",
                f"Você tem certeza que deseja excluir a conta '{selected_name}'?"
            )

            if confirm:
                for account in accounts:
                    if account["name"] == selected_name:
                        delete_account(account["name"])
                        messagebox.showinfo("Sucesso", "Conta excluída com sucesso!")
                        self.setup_accounts_tab()
                        popup.destroy()
                        return

        ctk.CTkButton(popup, text="Excluir", command=delete_selected).pack(pady=10)

        popup.protocol("WM_DELETE_WINDOW", popup.destroy)

    def setup_settings_tab(self):
        for widget in self.tab_settings.winfo_children():
            widget.destroy()

        # Centralizar elementos da aba de configurações
        settings_frame = ctk.CTkFrame(self.tab_settings)
        settings_frame.pack(expand=True, fill="both")

        manage_frame = ctk.CTkFrame(settings_frame)
        manage_frame.pack(pady=10)

        ctk.CTkLabel(manage_frame, text="Gerenciar Contas", font=("Arial", 14)).grid(row=0, column=0, columnspan=3, pady=5)

        add_icon = ctk.CTkButton(manage_frame, text="Adicionar", command=self.add_account, width=50)
        add_icon.grid(row=1, column=0, padx=5)

        edit_icon = ctk.CTkButton(manage_frame, text="Editar", command=self.edit_accounts_interface, width=50)
        edit_icon.grid(row=1, column=1, padx=5)

        delete_icon = ctk.CTkButton(manage_frame, text="Excluir", command=self.delete_account_interface, width=50)
        delete_icon.grid(row=1, column=2, padx=5)
        
        ctk.CTkLabel(settings_frame, text="Configurações", font=("Arial", 14)).pack(pady=5)

        coord_button = ctk.CTkButton(settings_frame, text="Coordenadas", command=capture_coordinates)
        coord_button.pack(pady=5)

        theme_button = ctk.CTkButton(settings_frame, text="Tema", command=self.toggle_theme)
        theme_button.pack(pady=5)

    def toggle_theme(self):
        current_theme = self.config.get("theme", "Dark")
        new_theme = "Light" if current_theme == "Dark" else "Dark"
        ctk.set_appearance_mode(new_theme)
        self.config["theme"] = new_theme
        save_config(self.config)

    def perform_login(self, account):
        config = load_config()
        if not config.get("account_coords") or not config.get("password_coords"):
            messagebox.showerror("Erro", "Coordenadas não configuradas.")
            return

        try:
            pyautogui.doubleClick(config["account_coords"])
            time.sleep(0.5)
            pyautogui.write(str(account["username"]))

            pyautogui.doubleClick(config["password_coords"])
            time.sleep(0.5)
            pyautogui.write(str(account["password"]))

            pyautogui.press("enter")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao realizar login: {e}")

if __name__ == "__main__":
    app = PasswordManagerApp()
    app.mainloop()
